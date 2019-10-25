# -*- coding:utf-8 -*-
import io
import os
from datetime import datetime
import time
import json

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


def readFileToList(path):
    '''
    读取txt存入list
    :param path: 文件路径
    :return: 列表
    '''
    with open(path) as fp:
        content = fp.readlines()
    list = [x.strip() for x in content]
    return list


def getChangeLogContent():
    '''
    获取到修改日志的内容
    :return:修改日志的内容list
    '''
    list = []
    title = ['', u'内容', 0]
    list.append(title)
    find_file_changelog = findFile('changelog')
    if find_file_changelog == None:
        return list
    content_list = readFileToList(find_file_changelog)
    list_add = []
    list_update = []
    list_fixbug = []
    list_other = []

    lenght = len(content_list)
    for i in range(0, lenght):
        item_content = content_list[i]
        if 'stop' in item_content:
            break
        if '新增' in item_content:
            list_add.append(item_content)
        elif '更新' in item_content:
            list_update.append(item_content)
        elif '修复' in item_content:
            list_fixbug.append(item_content)
        elif '其他' in item_content:
            list_other.append(item_content)

    content_add = [u'新增功能', getListContent(list_add), len(list_add)]
    content_udpate = [u'更新功能', getListContent(list_update), len(list_update)]
    content_fixbug = [u'修复bug', getListContent(list_fixbug), len(list_fixbug)]
    content_other = [u'其他', getListContent(list_other), len(list_other)]

    list.append(content_add)
    list.append(content_udpate)
    list.append(content_fixbug)
    list.append(content_other)

    return list


def getJsonObj(path):
    '''
    传入路径获取jsonc对象
    :param path:文件路径
    :return:对象
    '''

    try:
        # GBK
        file = io.open(path, 'r')
        obj = json.load(file)
        return obj
    except UnicodeDecodeError:
        # UTF-8
        with open(path, 'r') as f:
            data = json.load(f)
        return data


def getConfigContent():
    '''
    :return: 获取config文件的jso对象
    '''
    find_file_config = findFile('config')
    if find_file_config == None:
        return None
    json_obj = getJsonObj(find_file_config)
    return json_obj


def getAlgoContent():
    '''
    获取到算法表格的list内容
    :return:算法表格的list内容
    '''
    list = []
    title = [u'算法', u'说明', u'备注']
    list.append(title)

    find_file = findFile('algo')
    if find_file == None:
        return list

    json_obj = getJsonObj(find_file)
    # print json_obj
    json_arr_algo = json_obj['algo']
    # print json_arr_algo
    arrlen = len(json_arr_algo)
    for i in range(0, arrlen):
        name_ = json_arr_algo[i]['name']
        lib_ = json_arr_algo[i]['lib']
        tip_ = json_arr_algo[i]['tip']
        list.append([name_, lib_, tip_])
    return list


def getAddtionfileContent():
    '''
    获取到附加文件的表格内容
    :return:附加文件的表格内容list
    '''
    list = []
    title = [u'文件类型', u'大小', u'名称', u'md5']
    list.append(title)

    find_file_md5_addtionfile = findFile('md5_addtionfile')
    if find_file_md5_addtionfile == None:
        return list

    json_obj = getJsonObj(find_file_md5_addtionfile)
    json_arr_algo = json_obj['addtionfile']
    arrlen = len(json_arr_algo)
    for i in range(0, arrlen):
        type_ = json_arr_algo[i]['type']
        name_ = json_arr_algo[i]['name']
        md5_ = json_arr_algo[i]['md5']
        size_ = json_arr_algo[i]['size']
        if str(type_) == '1':
            type_ = u'软件包'
        elif str(type_) == '2':
            type_ = u'烧录软件'
        list.append([type_, size_, name_, md5_])

    find_file_testRepo = findFile('HardwareTestRepo')
    if find_file_testRepo == None:
        find_file_testRepo = ''
    addition_firmware_repo = [u'硬件报告', '', find_file_testRepo, '']
    list.append(addition_firmware_repo)
    addition_test_repo = [u'测试报告', '', '', '']
    list.append(addition_test_repo)
    addition_software_release_repo = [u'软件发布结论', '', '', '']
    list.append(addition_software_release_repo)
    return list


def getCheckListContent():
    '''
    获取到核对清单的表格内容
    :return:核对清单的表格内容 list
    '''
    list = []
    title = [u'事项', u'核对', u'签名', u'备注']
    list.append(title)

    find_file_checklist = findFile('checklist')
    json_obj = getJsonObj(find_file_checklist)
    json_arr_algo = json_obj['checklist']
    arrlen = len(json_arr_algo)
    for i in range(0, arrlen):
        list.append([json_arr_algo[i]['content'], '', '', ''])

    return list


def getReleaseInfo():
    '''
    获取到发布信息表格的内容
    :return:发布信息表格的内容list
    '''
    list = []

    title = [u'项目', u'版本号', u'SVN路径', u'发布文件路径']
    list.append(title)

    config = getConfigContent()
    if config == None:
        return list
    content_productType = config['productType']['content']
    content_firmwareVersion = config['firmwareVersion']['content']
    path_svn = config['svnPath']
    path_dir = config['dirPath']
    release_info = [content_productType, content_firmwareVersion, path_svn, path_dir]
    list.append(release_info)

    return list


def getOrderNumber():
    config = getConfigContent()
    if config == None:
        return "无"
    order_number = config['orderNumber']
    return order_number


def getProductInfoContent():
    '''
    获取到生产信息表格的内容.
    :return: 生产信息表格的内容list
    '''
    list = []

    title = [u'公司名称', u'产品型号', u'硬件版本', u'固件版本', u'生产日期', u'代理编号', u'生产批次']
    list.append(title)

    config = getConfigContent()
    if config == None:
        return list
    byte_companyName = config['companyName']['byte']
    byte_productType = config['productType']['byte']
    byte_hardwareVersion = config['hardwareVersion']['byte']
    byte_firmwareVersion = config['firmwareVersion']['byte']
    byte_productDate = config['productDate']['byte']
    byte_proxyNumber = config['proxyNumber']['byte']
    byte_productBatch = config['productBatch']['byte']
    product_info_1 = [byte_companyName, byte_productType, byte_hardwareVersion, byte_firmwareVersion, byte_productDate,
                      byte_proxyNumber, byte_productBatch]
    list.append(product_info_1)

    content_companyName = config['companyName']['content']
    content_productType = config['productType']['content']
    content_hardwareVersion = config['hardwareVersion']['content']
    content_firmwareVersion = config['firmwareVersion']['content']
    content_productDate = config['productDate']['content']
    content_proxyNumber = config['proxyNumber']['content']
    content_productBatch = config['productBatch']['content']
    product_info_2 = [content_companyName, content_productType, content_hardwareVersion, content_firmwareVersion,
                      content_productDate, content_proxyNumber, content_productBatch]
    list.append(product_info_2)

    return list


def getReleaseRecoderContent():
    '''
    获取到生产记录表格的内容
    :return: 生产记录表格的内容list
    '''
    list = []
    title = [u'时间', u'负责人', u'软件版本', u'硬件版本', u'协议', u'客户设备号', u'发布概要']
    list.append(title)
    config = getConfigContent()
    if config == None:
        return list
    date_ = datetime.now().strftime('%Y%m%d')
    person_ = config['person']
    content_firmwareVersion = config['firmwareVersion']['content']
    version_hardwareVersion = config['hardwareVersion']['content']
    version_protocol = config['protocolVersion']
    nubmer_content_ = config['deviceNumber']['content']
    batch = config['batch']
    deviceName = config['deviceName']
    detail_ = u'第' + batch + u'批,名称:' + deviceName
    release_recorder = [date_, person_, content_firmwareVersion, version_hardwareVersion, version_protocol,
                        nubmer_content_, detail_]
    list.append(release_recorder)
    return list


def getListContent(list):
    '''
    转换到修改日志的内容list->str
    :param list: 行内容
    :return: str字符串
    '''
    length = len(list)
    if length == 0:
        return ''
    content = '\n'
    for i in range(0, length):
        split_str = list[i].split('.')
        if len(split_str) >= 2:
            content = content + str(i + 1) + '.' + split_str[1] + '\n'
        else:
            content = content + str(i + 1) + '.' + list[i] + '\n'
    return content


def getProjectCheckContent():
    '''
    获取到项目审核表格的内容
    :return:项目审核表格的内容 list
    '''
    title = [u'', u'签名', u'时间', u'备注']
    content_1 = [u'项目审核人', '', '', '']
    content_2 = [u'硬件审核人', '', '', '']
    content_3 = [u'测试审核人', '', '', '']
    content_4 = [u'软件审核人', '', '', '']
    content_5 = [u'总监审核人', '', '', '']

    list = [title, content_1, content_2, content_3, content_4, content_5]
    return list


def getProjectMainInfo():
    '''
    获取项目信息概要表格的内容
    :return: 项目信息概要表格的内容list
    '''
    title = [u'产品型号', u'负责人', u'生产日期', u'软件版本', u'客户设备号', u'发布概要', u'备注']
    config = getConfigContent()
    if config == None:
        return list
    content_productType = config['productType']['content']
    content_person = config['person']
    content_productDate = config['productDate']['content']
    content_firmwareVersion = config['firmwareVersion']['content']
    content_device_number = config['deviceNumber']['content']

    batch = config['batch']
    deviceName = config['deviceName']
    content_detail = u'第' + batch + u'批,名称:' + deviceName

    content_1 = [content_productType, content_person, content_productDate, content_firmwareVersion,
                 content_device_number, content_detail,u'']

    list = [title, content_1]
    return list


def getProjectManagerContent():
    '''
    获取项目负责人表格的内容
    :return: 项目负责人表格的内容list
    '''
    title = [u'', u'签名', u'时间', u'备注']
    content_1 = [u'软件负责人', '', '', '']
    content_2 = [u'硬件负责人', '', '', '']
    content_3 = [u'测试负责人', '', '', '']

    list = [title, content_1, content_2, content_3]
    return list


def getRepoPersonAndTime():
    '''
    获取到发布人及日期的信息
    :return:  发布人及日期的信息str
    '''
    config = getConfigContent()
    u_time_ = u'  时间：' + datetime.now().strftime('%Y%m%d')
    if config == None:
        return u_time_
    person_ = config['person']
    return u'发布人：' + person_ + u_time_


def getFirstTitleContent():
    '''
    获取到表格的大标题 str
    :return:表格的大标题str
    '''
    u_title = u'软件版本发布'
    config = getConfigContent()
    if config == None:
        return u_title
    productType = config['productType']['content']
    proxy_name = config['proxyNumber']['name']
    return productType + '(' + proxy_name + ')' + u_title


def getFileName():
    '''
    获取到保存文件的文件名
    :return: 文件名
    '''
    config = getConfigContent()
    xlsx = 'productRelease.xlsx'
    if config == None:
        return xlsx
    productType = config['productType']['content']
    content_firmwareVersion = config['firmwareVersion']['content']
    nubmer_content_ = config['deviceNumber']['content']
    return productType + '_' + nubmer_content_ + '_' + content_firmwareVersion + '_' + xlsx


def findFile(keyword):
    '''
    根据关键字查找当前目录并返回文件路径
    :param keyword: 关键字
    :return: 文件路径
    '''
    rootdir = './'  # 需要遍历的文件夹，这里设定为当前文件夹
    list = os.listdir(rootdir)
    for line in list:
        if os.path.isfile(line):
            if keyword in line:
                return line
    return None


def block_first_title(ws):
    '''
    生成大标题
    :param ws:
    :return:占用的行数
    '''
    tileCell = ws['A1']
    tileCell.value = getFirstTitleContent()
    titlefont = Font(name=u'宋体',
                     size=font_size_one,
                     bold=True,
                     color='FF000000')
    tileCell.font = titlefont
    ws.merge_cells('A1:G1')
    mergerTitleCell = ws['A1']
    setCenter(mergerTitleCell)
    setCellHeight(1, ws, height_size_one)
    return 1


# sencond big Title:2,5,9..42
def gengerSecondTitle(ws, column, tileContent):
    '''
    生成二大标题
    :param ws:
    :param column:第几行
    :param tileContent:文本内容
    :return:
    '''
    setCellHeight(column, ws, height_size_two)
    column_start = 'A' + str(column)
    column_end = 'G' + str(column)
    cellRecoder = ws[column_start]
    cellRecoder.value = tileContent
    titlefont = Font(name=u'宋体',
                     size=font_size_two,
                     bold=True,
                     color='FF000000')
    cellRecoder.font = titlefont
    ws.merge_cells(column_start + ':' + column_end)
    mergerTitle2Cell = ws[column_start]
    setCenter(mergerTitle2Cell)


def gengerItemTitleContent(ws, column, tileContent, isTitle):
    '''
    生成一行内容,如第3行，及第6行的内容
    :param ws:
    :param column: 第几行
    :param tileContent: 文本内容 list
    :param isTitle: 加粗字体的判断
    :return:
    '''

    setThreeHeight(column, ws)
    column = str(column)
    len1 = len(tileContent)
    for i in range(65, 65 + len1):
        index = chr(i) + column
        cellsp = ws[index]
        cellsp.value = tileContent[i - 65]
        if isTitle == True:
            setBoldAndCenter(cellsp)
        else:
            setFont(cellsp)
            setCenter(cellsp)


def generTitileAndContent(ws, cloumnStart, column, title, isTitle):
    '''
    生成单个内容
    :param ws:
    :param cloumnStart: 第几列
    :param column: 第几行
    :param tileContent: 文本内容 str
    :param isTitle: 加粗字体的判断
    :return:
    '''
    setThreeHeight(column, ws)
    cloumn_start = cloumnStart + str(column)
    cellsp = ws[cloumn_start]
    cellsp.value = title
    if isTitle == True:
        setBoldAndCenter(cellsp)
    else:
        setFont(cellsp)
        setCenter(cellsp)


def generMegerTitileAndContent(ws, cloumnStart, cloumnEnd, column, title, isTitle):
    '''
    生成合并单元内容
    :param ws:
    :param cloumnStart: 第几列
    :param cloumnEnd: 第几列
    :param column: 第几行
    :param tileContent: 文本内容 str
    :param isTitle: 加粗字体的判断
    :return:
    '''
    setThreeHeight(column, ws)
    cloumn_start = cloumnStart + str(column)
    cloumn_end = cloumnEnd + str(column)
    cellsp = ws[cloumn_start]
    cellsp.value = title
    ws.merge_cells(cloumn_start + ':' + cloumn_end)
    cellsp = ws[cloumn_start]
    if isTitle == True:
        setBoldAndCenter(cellsp)
    else:
        setFont(cellsp)
        setCenter(cellsp)


def generMegerTitileAndContentHeight(ws, cloumnStart, cloumnEnd, column, title, isTitle, height):
    '''
    生成合并单元内容,并修改占用高度,仅用于修改记录表格
    :param ws:
    :param cloumnStart: 第几列开始
    :param cloumnEnd: 第几列结束
    :param column: 第几行
    :param tileContent: 文本内容 str
    :param isTitle: 加粗字体的判断
    :param height: 表格高度
    :return:
    '''
    setThreeHeight(column, ws)
    cloumn_start = cloumnStart + str(column)
    cloumn_end = cloumnEnd + str(column)
    cellsp = ws[cloumn_start]
    cellsp.value = title
    ws.merge_cells(cloumn_start + ':' + cloumn_end)
    cellsp = ws[cloumn_start]
    if isTitle == True:
        setBoldAndCenter(cellsp)
    else:
        setFont(cellsp)
        setLeft(cellsp)
    C_Column = ws.row_dimensions[column]
    C_Column.height = height


def setBorder(cell):
    '''
    单元格加边框
    :param cell:单元格
    :return:
    '''
    thin_border = Border(left=Side(style='thin', color='FF333333'),
                         right=Side(style='thin', color='FF333333'),
                         top=Side(style='thin', color='FF333333'),
                         bottom=Side(style='thin', color='FF333333'))
    cell.border = thin_border


def setCenter(cell):
    '''
    设置居中
    :param cell:单元格
    :return:
    '''
    align = Alignment(horizontal='center', vertical='center', wrap_text='True')
    cell.alignment = align


def setRight(cell):
    '''
    设置靠右
    :param cell:单元格
    :return:
    '''
    align = Alignment(horizontal='right', vertical='center', wrap_text='True')
    cell.alignment = align


def setLeft(cell):
    '''
    设置靠左
    :param cell:单元格
    :return:
    '''
    align = Alignment(horizontal='left', vertical='center', wrap_text='True')
    cell.alignment = align


def setFont(cell):
    '''
    设置字体加粗
    :param cell:单元格
    :return:
    '''
    titlefont = Font(name=u'宋体',
                     size=font_size_three,
                     color='FF000000')
    cell.font = titlefont


def setFontBold(cell):
    '''
    设置字体加粗
    :param cell:单元格
    :return:
    '''
    titlefont = Font(name=u'宋体',
                     size=font_size_three,
                     bold=True,
                     color='FF000000')
    cell.font = titlefont


def setBoldAndCenter(cell):
    '''
     设置字体加粗居中
     :param cell:单元格
     :return:
     '''
    setFontBold(cell)
    setCenter(cell)


def initConfig(ws):
    '''
    设置单元格的宽度
    :param ws:
    :return:
    '''
    for i in range(65, 72):
        A_Column = ws.column_dimensions[chr(i)]
        A_Column.width = 19


def setCellHeight(column, ws, height):
    '''
    设置单元格的高度
    :param column:行
    :param ws:
    :param height:高度
    :return:
    '''
    C_Column = ws.row_dimensions[column]
    C_Column.height = height


def setThreeHeight(column, ws):
    '''
    设置普通间单元格的高度
    :param column: 行
    :param ws:
    :return:
    '''
    setCellHeight(column, ws, height_size_three)


def block_release_recoder(ws, start_column, title):
    '''
    发布记录模块
    :param ws:
    :param start_column:开始行
    :param title: 标题
    :return:占用行
    '''
    sum_cloumn_release_recoder = 3
    gengerSecondTitle(ws, start_column, title)
    list = getReleaseRecoderContent()
    column_1 = start_column + 1
    for i in range(0, len(list)):
        isBold = True if (i == 0) else False
        gengerItemTitleContent(ws, column_1 + i, list[i], isBold)

    return sum_cloumn_release_recoder


def block_product_info(ws, start_column, title):
    '''
    生产信息模块
    :param ws:
    :param start_column:开始行
    :param title: 标题
    :return:占用行
    '''
    sum_cloumn_product_info = 5
    gengerSecondTitle(ws, start_column, title)

    # 临时插入一行订单编号
    column_1 = start_column + 1
    generMegerTitileAndContent(ws, 'A', 'C', column_1, u'订单编号', True)
    order_number = getOrderNumber()
    generMegerTitileAndContent(ws, 'D', 'G', column_1, order_number, False)

    column_1 = column_1 + 1
    list = getProductInfoContent()
    for i in range(0, len(list)):
        isBold = True if (i == 0) else False
        gengerItemTitleContent(ws, column_1 + i, list[i], isBold)
    return sum_cloumn_product_info


def block_release_info(ws, current_column, title):
    '''
    发布信息模块
    :param ws:
    :param start_column:开始行
    :param title: 标题
    :return:占用行
    '''
    sum_cloumn_release_info = 3
    gengerSecondTitle(ws, current_column, title)
    list = getReleaseInfo()
    column_1 = current_column + 1
    for i in range(0, len(list)):
        isBold = True if (i == 0) else False
        column_i = column_1 + i
        generTitileAndContent(ws, 'A', column_i, list[i][0], isBold)
        generTitileAndContent(ws, 'B', column_i, list[i][1], isBold)
        generMegerTitileAndContent(ws, 'C', 'D', column_i, list[i][2], isBold)
        generMegerTitileAndContent(ws, 'E', 'G', column_i, list[i][3], isBold)

    return sum_cloumn_release_info


def block_addtionfile_list(ws, current_column, title):
    '''
    附加文件模块
    :param ws:
    :param start_column:开始行
    :param title: 标题
    :return:占用行
    '''
    sum_cloumn_addtion_list = 7
    gengerSecondTitle(ws, current_column, title)

    list = getAddtionfileContent()
    column_1 = current_column + 1
    for i in range(0, len(list)):
        isBold = True if (i == 0) else False
        column_i = column_1 + i
        generTitileAndContent(ws, 'A', column_i, list[i][0], isBold)
        generTitileAndContent(ws, 'B', column_i, list[i][1], isBold)
        generMegerTitileAndContent(ws, 'C', 'E', column_i, list[i][2], isBold)
        generMegerTitileAndContent(ws, 'F', 'G', column_i, list[i][3], isBold)

    return sum_cloumn_addtion_list


def block_algo_version(ws, current_column, title):
    '''
    算法版本模块
    :param ws:
    :param start_column:开始行
    :param title: 标题
    :return:占用行
    '''
    sum_cloumn_algo_version = 9
    gengerSecondTitle(ws, current_column, title)
    list = getAlgoContent()
    column_1 = current_column + 1
    for i in range(0, len(list)):
        isBold = True if (i == 0) else False
        column_i = column_1 + i
        generMegerTitileAndContent(ws, 'A', 'B', column_i, list[i][0], isBold)
        generMegerTitileAndContent(ws, 'C', 'E', column_i, list[i][1], isBold)
        generMegerTitileAndContent(ws, 'F', 'G', column_i, list[i][2], isBold)

    return sum_cloumn_algo_version


def block_change_log(ws, current_column, title):
    '''
    修改日志模块
    :param ws:
    :param start_column:开始行
    :param title: 标题
    :return:占用行
    '''
    sum_cloumn_algo_changelog = 6
    gengerSecondTitle(ws, current_column, title)
    list = getChangeLogContent()
    column_1 = current_column + 1
    for i in range(0, len(list)):
        isBold = True if (i == 0) else False
        column_i = column_1 + i
        generTitileAndContent(ws, 'A', column_i, list[i][0], isBold)
        if isBold == False:
            height = (list[i][2] + 2) * height_size_log_oneline
            generMegerTitileAndContentHeight(ws, 'B', 'G', column_i, list[i][1], isBold, height)
        else:
            generMegerTitileAndContent(ws, 'B', 'G', column_i, list[i][1], isBold)
    return sum_cloumn_algo_changelog


def block_check_list(ws, current_column, title):
    '''
    核查列表模块
    :param ws:
    :param start_column:开始行
    :param title: 标题
    :return:占用行
    '''
    sum_cloumn_check_list = 1
    gengerSecondTitle(ws, current_column, title)
    list = getCheckListContent()
    current_column = current_column + 1
    sum_cloumn_check_list = sum_cloumn_check_list + len(list)
    for i in range(0, sum_cloumn_check_list - 1):
        isBold = True if (i == 0) else False
        current_column_i = current_column + i
        generMegerTitileAndContent(ws, 'A', 'B', current_column_i, list[i][0], isBold)
        generTitileAndContent(ws, 'C', current_column_i, list[i][1], isBold)
        generTitileAndContent(ws, 'D', current_column_i, list[i][2], isBold)
        generMegerTitileAndContent(ws, 'E', 'G', current_column_i, list[i][3], isBold)
    return sum_cloumn_check_list


def block_projec_main_info(ws, current_column, title):
    '''
    项目信息概要模块
    :param ws:
    :param start_column:开始行
    :param title: 标题
    :return:占用行
    '''
    sum_cloumn_projectmain_info = 3
    gengerSecondTitle(ws, current_column, title)
    list = getProjectMainInfo()
    current_column = current_column + 1
    for i in range(0, len(list)):
        isBold = True if (i == 0) else False
        gengerItemTitleContent(ws, current_column + i, list[i], isBold)
    return sum_cloumn_projectmain_info


def block_project_manager(ws, current_column, title):
    '''
    项目负责人模块
    :param ws:
    :param start_column:开始行
    :param title: 标题
    :return:占用行
    '''
    sum_cloumn_project_manager = 5
    gengerSecondTitle(ws, current_column, title)
    list = getProjectManagerContent()
    current_column = current_column + 1
    for i in range(0, len(list)):
        isBold = True if (i == 0) else False
        current_column_i = current_column + i
        generTitileAndContent(ws, 'A', current_column_i, list[i][0], isBold)
        generTitileAndContent(ws, 'B', current_column_i, list[i][1], isBold)
        generTitileAndContent(ws, 'C', current_column_i, list[i][2], isBold)
        generMegerTitileAndContent(ws, 'D', 'G', current_column_i, list[i][3], isBold)
    return sum_cloumn_project_manager


def block_project_check(ws, current_column, title):
    '''
    项目审核模块
    :param ws:
    :param start_column:开始行
    :param title: 标题
    :return:占用行
    '''
    sum_cloumn_project_check = 7
    gengerSecondTitle(ws, current_column, title)
    list = getProjectCheckContent()
    current_column = current_column + 1
    for i in range(0, len(list)):
        isBold = True if (i == 0) else False
        current_column_i = current_column + i
        generTitileAndContent(ws, 'A', current_column_i, list[i][0], isBold)
        generTitileAndContent(ws, 'B', current_column_i, list[i][1], isBold)
        generTitileAndContent(ws, 'C', current_column_i, list[i][2], isBold)
        generMegerTitileAndContent(ws, 'D', 'G', current_column_i, list[i][3], isBold)

    return sum_cloumn_project_check


def block_repo_person(ws, start_column):
    '''
    发布人及日期模块
    :param ws:
    :param start_column:开始行
    :param title: 标题
    :return:占用行
    '''
    sum_column = 1
    _time = getRepoPersonAndTime()
    generMegerTitileAndContent(ws, 'A', 'G', start_column, _time, False)
    cloumn_start = 'A' + str(start_column)
    cellsp = ws[cloumn_start]
    setRight(cellsp)
    return sum_column


def setAllBorder(sum_column):
    '''
    给所有单元格加上边框
    :param sum_column:
    :return:
    '''
    for column in range(2, sum_column):
        for i in range(65, 72):
            index = chr(i) + str(column)
            row = ws[index]
            setBorder(row)


# 一级字体的大小以及表格分配的高度
font_size_one = 45
height_size_one = 70

# 二级字体的大小以及表格分配的高度
font_size_two = 20
height_size_two = 43

# 三级字体的大小以及表格分配的高度
font_size_three = 16
height_size_three = 40

# 修改记录的一条信息的高度
height_size_log_oneline = 19

if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    ws.title = 'productRelease'
    initConfig(ws)
    # first title info
    sum_column = 1
    cloumn_count = block_first_title(ws)
    print u'排版大标题'
    # repo_person
    sum_column = sum_column + cloumn_count
    cloumn_count = block_repo_person(ws, sum_column)
    print u'排版备注'
    # recorder recoder
    sum_column = sum_column + cloumn_count
    cloumn_count = block_release_recoder(ws, sum_column, u'发布记录')
    print u'排版发布记录'
    # product info
    sum_column = sum_column + cloumn_count
    cloumn_count = block_product_info(ws, sum_column, u'生产信息')
    print u'排版生产信息'
    # release info
    sum_column = sum_column + cloumn_count
    cloumn_count = block_release_info(ws, sum_column, u'发布信息')
    print u'排版发布信息'
    # addition list
    sum_column = sum_column + cloumn_count
    cloumn_count = block_addtionfile_list(ws, sum_column, u'附件清单')
    print u'排版附件清单'
    # change log
    sum_column = sum_column + cloumn_count
    cloumn_count = block_change_log(ws, sum_column, u'修改记录')
    print u'排版修改记录'
    # algo version
    sum_column = sum_column + cloumn_count
    cloumn_count = block_algo_version(ws, sum_column, u'算法版本')
    print u'排版算法版本'
    # check list
    sum_column = sum_column + cloumn_count
    cloumn_count = block_check_list(ws, sum_column, u'注意事项核对')
    print u'排版注意事项核对'
    # project main info
    sum_column = sum_column + cloumn_count
    cloumn_count = block_projec_main_info(ws, sum_column, u'项目信息概要')
    print u'排版项目信息概要'
    # project manager
    sum_column = sum_column + cloumn_count
    cloumn_count = block_project_manager(ws, sum_column, u'项目负责')
    print u'排版项目负责'
    # project check
    sum_column = sum_column + cloumn_count
    cloumn_count = block_project_check(ws, sum_column, u'项目审核')
    print u'排版项目审核'
    # all border
    sum_column = sum_column + cloumn_count
    setAllBorder(sum_column)
    print u'排版边框处理'
    # save file
    file_name = getFileName()
    wb.save(file_name)
    print u'保存文件'
    time.sleep(5.5)
